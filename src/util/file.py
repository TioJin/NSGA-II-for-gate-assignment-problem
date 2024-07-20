"""文件相关工具

"""

import os
import openpyxl
from openpyxl import load_workbook, Workbook
from pathlib import Path

import pandas as pd


def create_folder(floder_path:str):
    try:
        path = Path(floder_path)
        path.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"|util -> file -> create_folder()| 创建文件时发生错误")
        raise e
    return True


def create_excel(file_path):
    """
        file_path: 文件的完整路径（包含文件名）
    """
    if os.path.exists(file_path):
        raise(Exception(f"excel文件已经存在 {file_path}，无法创建文件。")) # 如果外部存在了相同名字的文件则报错，以防止对该文件的数据造成损失。
    return openpyxl.Workbook()


class ExcelSave:

    # 各动作执行次数的表格名字
    action_sheet_name = 'actions'
    # 动作执行序列的表格名字
    action_sequence_sheet_name = 'action_sequence'
    excel = ".xlsx"

    def __init__(self) -> None:
        # Excel的存储路径
        self.save_path = None
        # Excel文件的名字
        self.file_name = None
        # Excel文件对象
        self.workbook: Workbook = None

    def _new_or_old_excel(self, save_path:str, file_name:str):
        """
        根据传进来的文件路径，判断是否创建新excel。
        """
        # 当传进来的路径与上一次一致时，认为用户需要在旧文件上进行追加，因此使用旧的excel文件对象。
        if save_path == self.save_path and file_name == self.file_name: 
            return self.workbook, False
        # 反之，认为用户需要使用新的excel文件，创建新的excel文件对象。
        else:
            self.save_path = save_path
            self.file_name = file_name
            create_folder(save_path)
            return create_excel(file_path=self.save_path + self.file_name), True

    def _new_or_old_sheet(self, workbook:Workbook, sheet_name:str):
        sheet = None
        """ 参数检查 """
        if type(workbook) != Workbook:
            raise(ValueError("传入的workbook参数不是Workbook类型数据"))
        
        """ 创建sheet """
        if sheet_name in workbook.sheetnames:
            # 旧sheet
            sheet = workbook[sheet_name]
        else:
            # 新sheet
            sheet = workbook.create_sheet(title=sheet_name)

        return sheet

    def _insert_data(self, sheet, data:list):
        """
        一行一行的插入数据。如果data是一个一维数组，则一个元素占excel一行，如果data是二位数组，
        则一个数组占excel一行。
        """
        for da in data:
            sheet.append(da)

    def save_arr_data(self, arr_data:list, save_path:str, file_name:str, sheet_name:str = "sheet"):
        """将二维数组中的内容保存至excel表中。数组的维度限制为一维和二维，数组内容限制为值类型。

        Parameters
        -----------
        arr_data: list
            输入数组。
        save_path: str
            保存的文件夹路径，例如：C:/Users/surface/OneDrive/Airport/Code/
        file_name: str
            保存的文件名字。
        sheet_name: str
            excel表格中的sheet名字。

        """
        self.workbook, flag = self._new_or_old_excel(save_path=save_path, file_name=file_name)
        sheet = self._new_or_old_sheet(self.workbook, sheet_name)
        # 如果excel表格是新建的，则删除表格中的默认sheet。
        if flag == True and len(self.workbook.sheetnames) >= 2:
            defalut_sheet = self.workbook["Sheet"]
            self.workbook.remove(defalut_sheet)
        self._insert_data(sheet, arr_data)
        final_path = save_path + file_name + ExcelSave.excel
        self.workbook.save(final_path)


def write_txt(file_path:str, file_name:str, content:str):
    """输入进来的路径必须是个文件夹路径，例如C:/Users/surface/OneDrive/Airport/Code/。
    """
    create_folder(file_path)
    file_path += file_name + ".txt"
    with open(file_path, "a") as file:
        file.write(content)


def read_excel(*args, **kwargs):
    """Read excel data.
    file_path, num_of_row.

    Parameters
    ----------
    The names of columns that you want to read. 
    file_path: str
        Use file_path to indicate the excel file.
    num_of_row: int
        How many rows you want to read from the excel file.

    Notes
    -----
    A row will be ignored if it has any missing value.
    """

    df = pd.read_excel(kwargs["file_path"], engine='openpyxl')
    df = df.fillna("BAD DATA")
    
    num_of_row = kwargs['num_of_row']
    if type(num_of_row) != int:
        raise(ValueError(f"读取的行的数量必须是int类型，但给出{type(num_of_row)}类型。"))

    excel_data = []
    for index, row in df.iterrows():
        if index == num_of_row: # 限制从文件中读取的行数。
            break
        row_data = []
        for arg in args:
            row_data.append(str(row[arg]))
        # 当前行存在空缺值，则忽略该行。
        if any(var == "BAD DATA" for var in row_data):
            continue
        excel_data.append(row_data)

    return excel_data
    

